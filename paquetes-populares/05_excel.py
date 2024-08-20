import openpyxl


wb = openpyxl.load_workbook("practica.xlsx")

# TODOS LOS CAMBIOS EJECUTADOS A UNA HOJA DE EXCEL SE EFECTUAN CUANDO GUARDAMOS DICHA HOJA COMO UN NUEVO PROYECTO(NO RESCRIBE LOS CAMBIOS EN LA MISMA HOJA)

# print(wb.sheetnames)
# print(wb["Sheet1"])
hoja = wb.active
# print(hoja)

# wb.create_sheet("Hoja 4")

# hoja4 = wb["Hoja 4"]
# hoja4.title = "Cambio de titulo"

# print(
#     hoja.max_row,
#     hoja.max_column
# )

# celda = hoja["B5"]
# celda.value = "nombres"
# print(celda.value)

# forma de acceder a los valores de las celdas de una forma mas programable(método cell)

# AQUI LOS INDICES PARTEN DESDE 1 NO DESDE CERO COMO ACOSTUMBRAN LAS ESTRUCTRAS DE DATOS
# celda2 = hoja.cell(row=6, column=2)
# print(
#     celda2.value,
#     celda2.row,
#     celda2.column,
#     celda2.coordinate
#     )

for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        # print(fila, columna, celda.value)

columna = hoja["A"]
fila = hoja["6"]
# print(fila)

# AGREGAR UNA FILA
hoja.append([243, "sam", 402])
# print(hoja.rows)

hoja.delete_rows(101, 8)

wb.save("practica1.xlsx")